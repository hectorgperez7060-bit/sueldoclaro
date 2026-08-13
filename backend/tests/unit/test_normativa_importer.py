from io import BytesIO

from openpyxl import load_workbook

from infrastructure.excel.normativa_importer import (
    COLUMNAS_NORMATIVAS,
    generar_plantilla_normativa,
    vista_previa_normativa,
)


def test_plantilla_tiene_columnas_e_instrucciones():
    contenido = generar_plantilla_normativa()
    wb = load_workbook(BytesIO(contenido), data_only=True)
    assert "Actualizacion normativa" in wb.sheetnames
    assert "Instrucciones" in wb.sheetnames
    encabezados = tuple(c.value for c in wb["Actualizacion normativa"][1])
    assert encabezados == COLUMNAS_NORMATIVAS


def test_preview_no_guarda_ni_habilita_aprobacion():
    res = vista_previa_normativa(generar_plantilla_normativa())
    assert res["total_filas"] == 3
    assert len(res["validas"]) == 3
    assert res["errores"] == []
    assert res["puede_aprobar"] is False
    assert "ninguna regla fue guardada" in res["mensaje"]


def test_preview_detecta_duplicado_y_fuente_faltante():
    wb = load_workbook(BytesIO(generar_plantilla_normativa()))
    ws = wb["Actualizacion normativa"]
    ws.append([c.value for c in ws[2]])
    ws.cell(row=5, column=COLUMNAS_NORMATIVAS.index("fuente") + 1).value = ""
    out = BytesIO(); wb.save(out)
    res = vista_previa_normativa(out.getvalue())
    assert len(res["errores"]) == 1
    assert "fila duplicada dentro del archivo" in res["errores"][0]["errores"]
    assert "fuente oficial obligatoria" in res["errores"][0]["errores"]
