"""Importación masiva de empleados desde .xlsx (openpyxl).

Valida fila por fila y reporta errores por fila sin abortar todo el archivo por
una fila mala (sección 5.2 del prompt).
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import List, Tuple

from openpyxl import Workbook, load_workbook

from domain.value_objects.cuil import digito_verificador, es_cuil_valido

from domain.entities.jornada import (
    HORAS_TOPE_LEY_11544,
    excede_limite_parcial,
    proporcion_jornada,
)
from infrastructure.excel.mapeo_columnas import (
    NOMBRE_COMPLETO,
    OBLIGATORIAS,
    SINONIMOS,
    detectar_mapeo,
    partir_nombre_completo,
)


def horas_jornada_de(horas_por_cct, cct_numero: str) -> Decimal:
    """Jornada completa del convenio de la fila; tope legal si no la declara."""
    valor = (horas_por_cct or {}).get(cct_numero)
    try:
        horas = Decimal(str(valor))
    except (ArithmeticError, TypeError, ValueError):
        return HORAS_TOPE_LEY_11544
    return horas if horas > 0 else HORAS_TOPE_LEY_11544


COLUMNAS = [
    "nombre", "apellido", "cuil", "fecha_ingreso", "cct_numero",
    "categoria", "legajo", "horas_semanales", "remuneracion_pactada",
    "afiliado_sindicato", "email",
]



def generar_cuil_valido_unico(dni_int: int, cuils_existentes: set[str], prefijo: str = "20") -> str:

    while True:
        diez_dig = f"{prefijo}{dni_int:08d}"
        dv = digito_verificador(diez_dig)
        if dv >= 0:
            c = f"{diez_dig}{dv}"
            if c not in cuils_existentes:
                return c
        dni_int += 1


def generar_plantilla(cuils_existentes: set[str] = None) -> bytes:
    cuils_existentes = cuils_existentes or set()
    cuil_sample = generar_cuil_valido_unico(12345678, cuils_existentes, "20")
    wb = Workbook()
    ws = wb.active
    ws.title = "empleados"
    ws.append(COLUMNAS)
    ws.append([
        "Juan", "Pérez", cuil_sample, "2021-07-01", "130/75",
        "Administrativo A", "0001", 48, "", "SI", "juan@ejemplo.com",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_demo_excel(cuils_existentes: set[str] = None) -> bytes:
    cuils_existentes = cuils_existentes or set()
    c1 = generar_cuil_valido_unico(12345678, cuils_existentes, "20")
    c2 = generar_cuil_valido_unico(23456789, cuils_existentes, "27")

    wb = Workbook()
    ws = wb.active
    ws.title = "empleados"
    ws.append(COLUMNAS)
    # Fila 2: Válida 1 (CUIL no existe en nómina activa)
    ws.append(["Carlos", "Gómez", c1, "2021-07-01", "130/75", "Administrativo A", "0001", 48, "", "SI", "carlos@ejemplo.com"])
    # Fila 3: Válida 2 (CUIL no existe en nómina activa)
    ws.append(["María", "López", c2, "2022-03-15", "130/75", "Vendedor B", "0002", 30, "", "NO", "maria@ejemplo.com"])
    # Fila 4: Error - CUIL repetido dentro del mismo Excel
    ws.append(["Carlos", "Gómez Duplicado", c1, "2021-07-01", "130/75", "Administrativo A", "0003", 48, "", "SI", ""])
    # Fila 5: Error - CUIL inválido
    ws.append(["Pedro", "Mendoza", "20999999999", "2023-01-10", "130/75", "Maestranza A", "0004", 48, "", "SI", ""])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()






FORMATOS_FECHA = (
    "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
    "%Y/%m/%d", "%d/%m/%y", "%d-%m-%y",
)


def _a_fecha(valor) -> date:
    """Acepta la fecha como la escriba el estudio, no sólo en ISO.

    En una planilla argentina la fecha de ingreso viene casi siempre como
    ``01/07/2021``. Exigir ``2021-07-01`` rechazaba archivos correctos.
    """
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    if not texto:
        raise ValueError("fecha vacía")
    if " " in texto:  # "01/07/2021 00:00:00"
        texto = texto.split(" ", 1)[0]
    for formato in FORMATOS_FECHA:
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    raise ValueError(f"formato de fecha no reconocido: {valor!r}")


def _a_texto_entero(valor) -> str:
    """Pasa a texto un número de Excel sin arrastrar el ``.0`` del float.

    Un CUIL o un legajo cargados como número llegan como ``20123456789.0`` y
    quedaban inválidos por culpa del formato, no del dato.
    """
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    if isinstance(valor, Decimal) and valor == valor.to_integral_value():
        return str(int(valor))
    return str(valor).strip()


def _a_decimal(valor) -> Decimal:
    """Interpreta importes escritos a la argentina: ``$ 1.234.567,89``."""
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    texto = str(valor).strip()
    texto = re.sub(r"[^0-9,.\-]", "", texto)
    if "," in texto and "." in texto:
        # El separador decimal es el último que aparece.
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", ".")
    if not texto or texto in ("-", ".", "-."):
        raise InvalidOperation("importe vacío")
    return Decimal(texto)


AFIRMATIVOS = {"SI", "S", "SÍ", "TRUE", "VERDADERO", "V", "1", "X", "Y", "YES", "AFILIADO"}


def _fila_encabezado(filas: List[tuple]) -> int:
    """Ubica la fila de títulos, salteando logos o títulos sueltos arriba.

    Muchas planillas traen el nombre de la empresa en A1 y recién en la fila 3
    los encabezados reales. Se elige la primera fila (de las cinco primeras) que
    reconozca al menos dos columnas conocidas.
    """
    mejor, mejor_puntaje = 0, -1
    for i, fila in enumerate(filas[:5]):
        indices, _, _ = detectar_mapeo(list(fila))
        puntaje = len(indices)
        if puntaje > mejor_puntaje:
            mejor, mejor_puntaje = i, puntaje
        if puntaje >= 4:
            break
    return mejor if mejor_puntaje >= 2 else 0


def _sugerencias(canonica: str) -> str:
    ejemplos = list(SINONIMOS.get(canonica, ()))[:4]
    return ", ".join(f'"{e}"' for e in ejemplos) if ejemplos else canonica


def parsear_con_mapeo(contenido: bytes, cuils_existentes: set[str] = None,
                      horas_por_cct: dict = None) -> Tuple[List[dict], List[dict], dict]:
    """Igual que :func:`parsear`, pero además explica cómo leyó el encabezado.

    El tercer valor es el informe de interpretación: qué columna del archivo se
    tomó para cada dato, cuáles se ignoraron y si hubo que partir una celda con
    apellido y nombre juntos. La UI lo muestra antes de confirmar la importación
    para que el usuario vea qué entendió la app.
    """
    wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    validos: List[dict] = []
    errores: List[dict] = []
    mapeo: dict = {"columnas": [], "ignoradas": [], "fila_encabezado": 1}
    if not filas:
        return validos, errores, mapeo

    fila_enc = _fila_encabezado(filas)
    encabezado_bruto = list(filas[fila_enc])
    idx, interpretacion, ignoradas = detectar_mapeo(encabezado_bruto)
    mapeo = {
        "columnas": interpretacion,
        "ignoradas": ignoradas,
        "fila_encabezado": fila_enc + 1,
    }

    # Una sola celda con apellido y nombre cubre ambas columnas obligatorias.
    usa_nombre_completo = NOMBRE_COMPLETO in idx and not ({"nombre", "apellido"} <= idx.keys())
    encabezado_nombre_completo = ""
    ancho = len(encabezado_bruto)
    if usa_nombre_completo:
        pos = idx[NOMBRE_COMPLETO]
        encabezado_nombre_completo = str(encabezado_bruto[pos] or "")
        mapeo["nombre_completo_partido"] = True
        # Dos posiciones virtuales al final de cada fila para el nombre partido.
        idx["nombre"], idx["apellido"] = ancho, ancho + 1

    faltantes = [
        c for c in OBLIGATORIAS
        if c not in idx and not (usa_nombre_completo and c in ("nombre", "apellido"))
    ]
    if faltantes:
        detalle = [
            f"No encontré la columna «{c}». Podés titularla {_sugerencias(c)}."
            for c in faltantes
        ]
        reconocidas = ", ".join(i["columna_archivo"] for i in interpretacion) or "ninguna"
        detalle.append(f"Columnas que sí reconocí: {reconocidas}.")
        return validos, [{"fila": fila_enc + 1, "errores": detalle}], mapeo

    cuils_vistos_excel: set[str] = set()
    cuils_existentes = cuils_existentes or set()

    for n, fila_origen in enumerate(filas[fila_enc + 1:], start=fila_enc + 2):
        if all(c is None or str(c).strip() == "" for c in fila_origen):
            continue  # fila en blanco al pie de la planilla

        fila = list(fila_origen) + [None] * (ancho - len(fila_origen))
        if usa_nombre_completo:
            _ape, _nom = partir_nombre_completo(fila[pos], encabezado_nombre_completo)
            fila = fila + [_nom, _ape]

        def val(col, _fila=fila):
            i = idx.get(col)
            return _fila[i] if (i is not None and i < len(_fila)) else None

        errs: List[str] = []
        cuil = _a_texto_entero(val("cuil")).replace("-", "").replace(" ", "").strip()
        if not es_cuil_valido(cuil):
            errs.append(f"CUIL inválido: {val('cuil')!r}")
        else:
            if cuil in cuils_vistos_excel:
                errs.append(f"CUIL {cuil} repetido dentro del mismo archivo Excel")
            else:
                cuils_vistos_excel.add(cuil)

            if cuil in cuils_existentes:
                errs.append(f"CUIL {cuil} ya existe registrado en la nómina activa")

        try:
            fecha_ing = _a_fecha(val("fecha_ingreso"))
        except (ValueError, TypeError):
            fecha_ing = None
            errs.append(
                f"fecha_ingreso inválida: {val('fecha_ingreso')!r} "
                "(se aceptan 01/07/2021, 01-07-2021 o 2021-07-01)"
            )

        for req in ("nombre", "apellido", "cct_numero", "categoria"):
            if not str(val(req) or "").strip():
                errs.append(f"Campo obligatorio vacío: {req}")

        rem = val("remuneracion_pactada")
        rem_dec = None
        if rem not in (None, ""):
            try:
                rem_dec = _a_decimal(rem)
            except (InvalidOperation, ValueError):
                errs.append(f"remuneracion_pactada inválida: {rem!r}")

        cct_fila = str(val("cct_numero") or "").strip()
        horas_completas = horas_jornada_de(horas_por_cct, cct_fila)
        try:
            crudo_horas = val("horas_semanales")
            horas_semanales = (
                _a_decimal(crudo_horas)
                if crudo_horas not in (None, "") else horas_completas
            )
            proporcion = proporcion_jornada(horas_semanales, horas_completas)
        except (InvalidOperation, ValueError) as exc:
            horas_semanales = horas_completas
            proporcion = Decimal("1")
            errs.append(str(exc) if isinstance(exc, ValueError) and str(exc)
                        else f"horas_semanales inválidas para el convenio {cct_fila}")

        # Mejor avisarlo acá que dejar un legajo que recién falla al liquidar.
        if excede_limite_parcial(proporcion):
            pct = (proporcion * 100).quantize(Decimal("0.01"))
            errs.append(
                f"{horas_semanales} horas semanales son el {pct}% de la jornada del "
                f"convenio {cct_fila} ({horas_completas} h), y eso supera las dos "
                "terceras partes. Por el art. 92 ter de la LCT le corresponde el "
                "sueldo de jornada completa: dejá la columna de horas vacía para "
                "cargarlo a jornada completa, o corregí las horas si trabaja menos."
            )

        nom = str(val("nombre") or "").strip()
        ape = str(val("apellido") or "").strip()

        if errs:
            errores.append({"fila": n, "nombre": f"{ape}, {nom}".strip(", "), "cuil": cuil, "errores": errs})
            continue

        crudo_afil = val("afiliado_sindicato")
        if isinstance(crudo_afil, bool):
            afiliado = crudo_afil
        else:
            afil = str(crudo_afil if crudo_afil not in (None, "") else "SI").strip().upper()
            afiliado = afil in AFIRMATIVOS
        validos.append({
            "fila": n,
            "nombre": nom,
            "apellido": ape,
            "cuil": cuil,
            "fecha_ingreso": fecha_ing,
            "cct_numero": str(val("cct_numero")).strip(),
            "categoria": str(val("categoria")).strip(),
            "legajo": _a_texto_entero(val("legajo")).strip(),
            "remuneracion_pactada": rem_dec,
            "proporcion_jornada": proporcion,
            "afiliado_sindicato": afiliado,
            "email": (str(val("email")).strip() or None) if val("email") else None,
        })

    return validos, errores, mapeo


def parsear(contenido: bytes, cuils_existentes: set[str] = None,
            horas_por_cct: dict = None) -> Tuple[List[dict], List[dict]]:
    """Devuelve (empleados_validos, errores_por_fila).

    ``horas_por_cct`` trae las horas de jornada completa declaradas por cada
    convenio. La proporción de jornada se calcula contra la jornada del convenio
    de esa fila, no contra 48 fijas: en un convenio de 44 horas, 44 horas son
    jornada completa y prorratear contra 48 le recortaría el sueldo.
    """
    validos, errores, _ = parsear_con_mapeo(contenido, cuils_existentes, horas_por_cct)
    return validos, errores
