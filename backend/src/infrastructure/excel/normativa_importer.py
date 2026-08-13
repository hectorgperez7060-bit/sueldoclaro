"""Plantilla y vista previa normativa. Nunca modifica reglas vigentes."""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook


COLUMNAS_NORMATIVAS = (
    "tipo", "cct_numero", "codigo", "categoria", "valor", "unidad", "ambito",
    "vigencia_desde", "vigencia_hasta", "fuente", "integra_antiguedad",
    "integra_presentismo", "aporte_jubilacion", "aporte_obra_social",
    "aporte_sindicato",
)
TIPOS = {"escala", "concepto", "deduccion"}
UNIDADES = {"ARS", "%"}
AMBITOS = {"no_rem", "ded_todos", "ded_afil", "ded_noafil"}


def generar_plantilla_normativa() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Actualizacion normativa"
    ws.append(COLUMNAS_NORMATIVAS)
    ws.append([
        "escala", "130/75", "", "Maestranza A", "0", "ARS", "", "2026-08-01",
        "", "ADJUNTAR FUENTE OFICIAL", "", "", "", "", "",
    ])
    ws.append([
        "concepto", "130/75", "COMERCIO_NR_EJEMPLO", "", "0", "ARS", "no_rem",
        "2026-08-01", "", "ADJUNTAR FUENTE OFICIAL", "SI", "SI", "NO", "SI", "SI",
    ])
    ws.append([
        "deduccion", "130/75", "APORTE_GREMIAL_EJEMPLO", "", "0.02", "%",
        "ded_todos", "2026-08-01", "", "ADJUNTAR FUENTE OFICIAL", "", "", "", "", "",
    ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            36, max(len(str(c.value or "")) for c in col) + 2
        )
    info = wb.create_sheet("Instrucciones")
    info.append(["Esta plantilla sólo prepara una vista previa. No aprueba ni publica reglas."])
    info.append(["tipo", "escala, concepto o deduccion"])
    info.append(["valor", "Importe ARS o fracción porcentual: 0.02 equivale a 2%"])
    info.append(["fuente", "Acuerdo, resolución, circular o escala oficial obligatoria"])
    info.append(["incidencias", "Usar SI/NO únicamente para conceptos en ARS"])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _fecha(valor: Any) -> date | None:
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return date.fromisoformat(str(valor).strip())


def _si_no(valor: Any, nombre: str, errores: list[str]) -> bool:
    normalizado = str(valor or "").strip().upper()
    if normalizado not in {"SI", "NO"}:
        errores.append(f"{nombre} debe indicar SI o NO")
        return False
    return normalizado == "SI"


def vista_previa_normativa(contenido: bytes) -> dict:
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError("El archivo no es un Excel .xlsx válido") from exc
    ws = wb["Actualizacion normativa"] if "Actualizacion normativa" in wb.sheetnames else wb.active
    filas = ws.iter_rows(values_only=True)
    encabezados = tuple(str(v or "").strip() for v in next(filas, ()))
    faltantes = [c for c in COLUMNAS_NORMATIVAS if c not in encabezados]
    if faltantes:
        raise ValueError("Faltan columnas obligatorias: " + ", ".join(faltantes))
    indices = {nombre: encabezados.index(nombre) for nombre in COLUMNAS_NORMATIVAS}
    validas, errores_salida, claves = [], [], set()

    for numero_fila, valores in enumerate(filas, start=2):
        if not any(v not in (None, "") for v in valores):
            continue
        d = {c: valores[i] if i < len(valores) else None for c, i in indices.items()}
        errores: list[str] = []
        tipo = str(d["tipo"] or "").strip().lower()
        cct = str(d["cct_numero"] or "").strip()
        codigo = str(d["codigo"] or "").strip()
        categoria = str(d["categoria"] or "").strip()
        unidad = str(d["unidad"] or "").strip().upper()
        ambito = str(d["ambito"] or "").strip()
        fuente = str(d["fuente"] or "").strip()
        if tipo not in TIPOS:
            errores.append("tipo inválido")
        if not cct:
            errores.append("cct_numero es obligatorio")
        if tipo == "escala" and not categoria:
            errores.append("la escala requiere categoría")
        if tipo != "escala" and not codigo:
            errores.append("el concepto requiere código")
        if unidad not in UNIDADES:
            errores.append("unidad debe ser ARS o %")
        if tipo in {"concepto", "deduccion"} and ambito not in AMBITOS:
            errores.append("ámbito inválido")
        if not fuente:
            errores.append("fuente oficial obligatoria")
        try:
            valor = Decimal(str(d["valor"]))
            if valor < 0:
                errores.append("valor no puede ser negativo")
        except (InvalidOperation, TypeError):
            valor = Decimal("0")
            errores.append("valor inválido")
        try:
            desde, hasta = _fecha(d["vigencia_desde"]), _fecha(d["vigencia_hasta"])
            if desde is None:
                errores.append("vigencia_desde obligatoria")
            if desde and hasta and hasta < desde:
                errores.append("vigencia_hasta anterior a vigencia_desde")
        except ValueError:
            desde = hasta = None
            errores.append("vigencia inválida; usar AAAA-MM-DD")

        clave = (tipo, cct, categoria if tipo == "escala" else codigo, desde)
        if clave in claves:
            errores.append("fila duplicada dentro del archivo")
        claves.add(clave)
        incidencias = {}
        if tipo == "concepto":
            for campo in COLUMNAS_NORMATIVAS[10:]:
                incidencias[campo] = _si_no(d[campo], campo, errores)

        normalizada = {
            "fila": numero_fila, "tipo": tipo, "cct_numero": cct,
            "codigo": codigo, "categoria": categoria, "valor": str(valor),
            "unidad": unidad, "ambito": ambito,
            "vigencia_desde": desde.isoformat() if desde else None,
            "vigencia_hasta": hasta.isoformat() if hasta else None,
            "fuente": fuente, "incidencias": incidencias,
        }
        if errores:
            errores_salida.append({**normalizada, "errores": errores})
        else:
            validas.append(normalizada)
    return {
        "total_filas": len(validas) + len(errores_salida),
        "validas": validas,
        "errores": errores_salida,
        "puede_aprobar": False,
        "mensaje": "Vista previa solamente: ninguna regla fue guardada ni aprobada",
    }
